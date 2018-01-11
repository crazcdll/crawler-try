import os
from openpyxl import Workbook

# 存储目录
o_path = os.getcwd()
file_path = o_path + '/data/'

wb = Workbook()
ws = wb.active
ws['A1'] = 42
ws.append([1, 2, 3])
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save(file_path + "/sample.xlsx")

with open(file_path + 'weather_test.txt', 'wt') as f:
    print('Hello World', file=f)

20140324
20140808
20140822


2009
2010
2011
2012
2013
2014
2015
2016
2017


<tr>
<th>Time <span class="nb">(CST)</span></th>
<th>Temp.</th>
<th>Windchill</th>
<th>Dew Point</th>
<th>Humidity</th>
<th>Pressure</th>
<th>Visibility</th>
<th>Wind Dir</th>
<th>Wind Speed</th>
<th>Gust Speed</th>
<th>Precip</th>
<th>Events</th>
<th>Conditions</th>
</tr>
</thead>
