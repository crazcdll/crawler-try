import scrapy
import os
from openpyxl import Workbook

# 存储目录
o_path = os.getcwd()
file_path = o_path + '/data/'

# 年份
years = ['2014', '2015', '2016', '2017']

# 月份
months = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12']

# 初始化 openpyxl 的 Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 'days'
ws['B1'] = 'AQI'
ws['C1'] = 'PM2.5'
ws['D1'] = 'PM10'
ws['E1'] = 'SO2'
ws['F1'] = 'NO2'
ws['G1'] = 'CO'
ws['H1'] = 'O3'


class AqiSpider(scrapy.Spider):
    name = 'aqispider'
    start_urls = ['http://www.tianqihoubao.com/aqi/beijing-201801.html']

    def parse(self, response):

        for tr in response.css('table.b tr'):
            row = []
            for index, td in enumerate(tr.css('td ::text')):
                if index == 0:
                    row.append(str(td.extract()).strip().replace('-', ''))
                    index += 1
                    continue
                if index == 2:
                    row.append(str(td.extract()).strip())
                    index += 1
                    continue
                if index == 4:
                    row.append(str(td.extract()).strip())
                    index += 1
                    continue
                if index == 5:
                    row.append(str(td.extract()).strip())
                    index += 1
                    continue
                if index == 6:
                    row.append(str(td.extract()).strip())
                    index += 1
                    continue
                if index == 7:
                    row.append(str(td.extract()).strip())
                    index += 1
                    continue
                if index == 8:
                    row.append(str(td.extract()).strip())
                    index += 1
                    continue
                if index == 9:
                    row.append(str(td.extract()).strip())
                    index += 1
                    continue
            ws.append(row)

        for next_year in years:
            for next_month in months:
                yield response.follow('http://www.tianqihoubao.com/aqi/beijing-' + next_year + next_month + '.html',
                                      self.parse)

        wb.save(file_path + "/aqi.xlsx")
