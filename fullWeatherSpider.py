import scrapy
import os
from openpyxl import Workbook

# 存储目录
o_path = os.getcwd()
file_path = o_path + '/data/'

# 年份
# years = ['1996', '1997', '1998', '1999', '2000', '2001', '2002', '2003', '2004', '2005', '2006', '2007', '2008', '2009',
#          '2010', '2011', '2012', '2013', '2014', '2015', '2016', '2017']
years = ['2017']

# 月份
months = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']

# 英文月份
eng_months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

wb = Workbook()
ws = wb.active

ws['A1'] = 'days'
ws['B1'] = 'temp-high'
ws['C1'] = 'temp-avg'
ws['D1'] = 'temp-low'
ws['E1'] = 'dew-point-temp-high'
ws['F1'] = 'dew-point-temp-avg'
ws['G1'] = 'dew-point-temp-low'
ws['H1'] = 'rh-high'
ws['I1'] = 'rh-avg'
ws['J1'] = 'rh-low'
ws['K1'] = 'sea-level-press-high'
ws['L1'] = 'sea-level-press-avg'
ws['M1'] = 'sea-level-press-low'
ws['N1'] = 'visibility-high'
ws['O1'] = 'visibility-avg'
ws['P1'] = 'visibility-low'
ws['Q1'] = 'wind-high'
ws['R1'] = 'wind-avg'
ws['S1'] = 'wind-low'
ws['T1'] = 'precip'
ws['U1'] = 'activity'

# 清空文件
# f = open(file_path + 'weather_test.txt', 'w')
# f.truncate();
# f.close()


class WeatherSpider(scrapy.Spider):
    name = 'weatherspider'
    start_urls = [
        'https://www.wunderground.com/history/airport/ZBAA/2018/1/1/MonthlyHistory.html?req_city=Beijing&req_statename=China&reqdb.zip=&reqdb.magic=&reqdb.wmo=']

    def parse(self, response):

        # 获得年份
        year = response.css('table.daily thead tr > th ::text').extract_first()

        # 获得月份
        month = response.css('table.daily tbody tr td ::text').extract_first()
        for index, eng_month in enumerate(eng_months):
            if month == eng_month:
                month = str(index + 1)
                if int(month) < 10:
                    month = '0' + str(month)

        for tbody in response.css('table.daily tbody'):
            row = []
            if tbody.css('tr > td').extract()[0].find('href') != -1:
                for index, td in enumerate(tbody.css('tr td')):
                    if index == 0:
                        day = ''.join(td.css('a ::text').extract())
                        if int(day) < 10:
                            day = '0' + day
                        row.append(year + month + day)
                        index += 1
                        continue
                    else:
                        row.append(''.join(td.css('span ::text').extract()))
                        index += 1
                        continue
                    # if index == 1:
                    #     row.append(''.join(td.css('span ::text').extract()))
                    #     index += 1
                    #     continue
                    # if index == 2:
                    #     row.append(''.join(td.css('span ::text').extract()))
                    #     index += 1
                    #     continue
                    # if index == 3:
                    #     row.append(''.join(td.css('span ::text').extract()))
                    #     index += 1
                    #     continue
                    # if index == 4:
                    #     row.append(''.join(td.css('span ::text').extract()))
                    #     index += 1
                    #     continue
                    # if index == 5:
                    #     row.append(''.join(td.css('span ::text').extract()))
                    #     index += 1
                    #     continue
                    # if index == 6:
                    #     row.append(''.join(td.css('span ::text').extract()))
                    #     index += 1
                    #     continue
                    # if index == 7:
                    #     row.append(''.join(td.css('span ::text').extract()))
                    #     index += 1
                    #     continue
                    # if index == 8:
                    #     row.append(''.join(td.css('span ::text').extract()))
                    #     index += 1
                    #     continue
                    # if index == 9:
                    #     row.append(''.join(td.css('span ::text').extract()))
                    #     index += 1
                    #     continue
                # with open(file_path + 'weather_test.txt', 'a') as f:
                #     print(row, file=f)
                ws.append(row)

        for next_year in years:
            for next_month in months:
                # https://www.wunderground.com/history/airport/ZBAA/2018/1/1/MonthlyHistory.html?req_city=Beijing&req_statename=China&reqdb.zip=&reqdb.magic=&reqdb.wmo=
                yield response.follow(
                    'https://www.wunderground.com/history/airport/ZBAA/' + next_year + '/' + next_month + '/1/MonthlyHistory.html?req_city=Beijing&req_statename=China&reqdb.zip=&reqdb.magic=&reqdb.wmo=',
                    self.parse)
        # 保存文件
        wb.save(file_path + "/weather.xlsx")
